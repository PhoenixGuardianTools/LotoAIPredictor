import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.database import get_game_stats, get_jackpot, get_special_days
import glob
import os

def rotate_reports(report_dir, max_reports=5):
    """Gère la rotation des rapports en gardant uniquement les plus récents."""
    # Créer le dossier s'il n'existe pas
    Path(report_dir).mkdir(parents=True, exist_ok=True)
    
    # Lister tous les rapports
    reports = glob.glob(os.path.join(report_dir, "*"))
    
    # Trier par date de modification (plus récent en premier)
    reports.sort(key=os.path.getmtime, reverse=True)
    
    # Supprimer les anciens rapports
    for old_report in reports[max_reports:]:
        try:
            os.remove(old_report)
        except Exception as e:
            print(f"Erreur lors de la suppression de {old_report}: {e}")

def get_special_day_info(date):
    """Récupère les informations sur les jours spéciaux."""
    special_days = get_special_days()
    date_obj = datetime.strptime(date, '%Y-%m-%d')
    for day in special_days:
        if day['date'] == date:
            return day['name'], day['description']
    return None, None

def export_to_excel(report, report_type, output_path=None):
    """
    Exporte un rapport au format Excel avec tris, macros et graphiques.
    
    Args:
        report (dict): Le rapport à exporter
        report_type (str): Type de rapport ('daily', 'weekly', 'monthly')
        output_path (str, optional): Chemin de sortie du fichier Excel
    """
    if output_path is None:
        output_path = f"reports/excel/{report['jeu']}_{report['date'] or report['mois']}.xlsx"
    
    # Gérer la rotation des rapports
    rotate_reports(os.path.dirname(output_path))
    
    # Créer le dossier de sortie si nécessaire
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Créer un writer Excel
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Feuille Résumé
    summary_data = {
        'Métrique': [
            'Jeu',
            'Date du tirage',
            'Jour du tirage',
            'Cagnotte actuelle',
            'Gain brut total',
            'Coût investi',
            'Gain net',
            'Gain net cumulé',
            'Ratio de gain',
            'Indice de confiance global'
        ],
        'Valeur': [
            report['jeu'],
            report['date'],
            report['jour_tirage'],
            f"{report['cagnotte']:,.2f} €",
            f"{report['gains_bruts']:,.2f} €",
            f"{report['cout_investi']:,.2f} €",
            f"{report['gain_net']:,.2f} €",
            f"{report['gain_net_cumule']:,.2f} €",
            f"{report['ratio_gain']:.2%}",
            f"{report['indice_confiance_global']:.2f}"
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Résumé', index=False)
    
    # Feuille Prédictions
    preds_data = []
    for pred in report['predictions']:
        pred_data = {
            'Numéros': ', '.join(map(str, pred['numbers'])),
            'Spéciaux': ', '.join(map(str, pred['special'])),
            'Score': f"{pred['score']:.2f}",
            'Indice Confiance': f"{pred['indice_confiance']:.2f}",
            'Gain Estimé': f"{pred['gain_estime']:,.2f} €",
            'Jackpot': '🍀' if pred.get('jackpot_predit', False) else ''
        }
        preds_data.append(pred_data)
    preds_df = pd.DataFrame(preds_data)
    preds_df.to_excel(writer, sheet_name='Prédictions', index=False)
    
    # Feuille Gains Prédits
    gains_predits_data = {
        'Rang': report['gains_predits']['rangs'],
        'Gain Estimé': [f"{g:,.2f} €" for g in report['gains_predits']['gains']],
        'Probabilité': [f"{p:.2%}" for p in report['gains_predits']['probabilites']]
    }
    gains_predits_df = pd.DataFrame(gains_predits_data)
    gains_predits_df.to_excel(writer, sheet_name='Gains Prédits', index=False)
    
    # Feuille Historique
    historique_df = pd.DataFrame(report['historique'])
    historique_df.to_excel(writer, sheet_name='Historique', index=False)
    
    # Obtenir le workbook pour les modifications
    workbook = writer.book
    
    # Appliquer le style à chaque feuille
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    # Ajouter des graphiques
    if 'historique' in report:
        # Graphique d'évolution des gains
        gains_chart = LineChart()
        gains_chart.title = "Évolution des gains"
        gains_chart.style = 10
        gains_chart.y_axis.title = "Gains (€)"
        gains_chart.x_axis.title = "Date"
        
        data = Reference(worksheet=workbook['Historique'], 
                        min_row=1, max_row=len(report['historique'])+1,
                        min_col=2, max_col=2)
        cats = Reference(worksheet=workbook['Historique'],
                        min_row=2, max_row=len(report['historique'])+1,
                        min_col=1, max_col=1)
        
        gains_chart.add_data(data, titles_from_data=True)
        gains_chart.set_categories(cats)
        
        workbook['Historique'].add_chart(gains_chart, "F2")
    
    # Sauvegarder le fichier
    writer.close()
    return output_path

def export_to_pdf(report, report_type, output_path=None):
    """
    Exporte un rapport au format PDF.
    
    Args:
        report (dict): Le rapport à exporter
        report_type (str): Type de rapport ('daily', 'weekly', 'monthly')
        output_path (str, optional): Chemin de sortie du fichier PDF
    """
    if output_path is None:
        output_path = f"reports/pdf/{report['jeu']}_{report['date'] or report['mois']}.pdf"
    
    # Gérer la rotation des rapports
    rotate_reports(os.path.dirname(output_path))
    
    # Créer le dossier de sortie si nécessaire
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Créer le document PDF
    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.green,
        spaceAfter=30
    )
    
    # Vérifier si c'est un jour spécial
    special_day, special_desc = get_special_day_info(report['date'])
    if special_day:
        title_text = f"Rapport {report_type} - {report['jeu']} - {special_day}"
    else:
        title_text = f"Rapport {report_type} - {report['jeu']}"
    
    elements.append(Paragraph(title_text, title_style))
    
    # Date du tirage
    date_style = ParagraphStyle(
        'CustomDate',
        parent=styles['Heading2'],
        fontSize=18,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Tirage de {report['jour_tirage']} {report['date']}", date_style))
    
    if special_day:
        elements.append(Paragraph(special_desc, styles['Normal']))
    
    elements.append(Spacer(1, 20))
    
    # Résumé
    elements.append(Paragraph("Résumé", styles['Heading2']))
    summary_data = [
        ['Métrique', 'Valeur'],
        ['Cagnotte actuelle', f"{report['cagnotte']:,.2f} €"],
        ['Gain brut total', f"{report['gains_bruts']:,.2f} €"],
        ['Coût investi', f"{report['cout_investi']:,.2f} €"],
        ['Gain net', f"{report['gain_net']:,.2f} €"],
        ['Gain net cumulé', f"{report['gain_net_cumule']:,.2f} €"],
        ['Ratio de gain', f"{report['ratio_gain']:.2%}"],
        ['Indice de confiance global', f"{report['indice_confiance_global']:.2f}"]
    ]
    summary_table = Table(summary_data, colWidths=[200, 200])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Prédictions
    elements.append(Paragraph("Prédictions", styles['Heading2']))
    preds_data = [['Numéros', 'Spéciaux', 'Score', 'Confiance', 'Gain Estimé', 'Jackpot']]
    for pred in report['predictions']:
        preds_data.append([
            ', '.join(map(str, pred['numbers'])),
            ', '.join(map(str, pred['special'])),
            f"{pred['score']:.2f}",
            f"{pred['indice_confiance']:.2f}",
            f"{pred['gain_estime']:,.2f} €",
            '🍀' if pred.get('jackpot_predit', False) else ''
        ])
    preds_table = Table(preds_data, colWidths=[120, 80, 60, 60, 80, 40])
    preds_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(preds_table)
    elements.append(Spacer(1, 20))
    
    # Gains Prédits
    elements.append(Paragraph("Gains Prédits", styles['Heading2']))
    gains_data = [['Rang', 'Gain Estimé', 'Probabilité']]
    for i in range(len(report['gains_predits']['rangs'])):
        gains_data.append([
            report['gains_predits']['rangs'][i],
            f"{report['gains_predits']['gains'][i]:,.2f} €",
            f"{report['gains_predits']['probabilites'][i]:.2%}"
        ])
    gains_table = Table(gains_data, colWidths=[100, 150, 150])
    gains_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(gains_table)
    
    # Pied de page
    footer_style = ParagraphStyle(
        'CustomFooter',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey
    )
    elements.append(Paragraph(
        f"Rapport généré le {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        footer_style
    ))
    
    # Construire le PDF
    doc.build(elements)
    return output_path 